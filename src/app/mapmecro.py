from bs4 import BeautifulSoup
import openpyxl as excel, urllib.parse, time, datetime, os
from PyQt5 import QtCore
import traceback, random, requests, json, logging, pickle

class Mecro(QtCore.QObject):
    percentChanged = QtCore.pyqtSignal(int)
    completion = QtCore.pyqtSignal(list)
    permission = QtCore.pyqtSignal(int)
    waitsignal = QtCore.pyqtSignal(int)

    def __init__(self, log_address):
        super().__init__()
        self.online = False
        self.logger = logging.getLogger('Mecro Logger')
        self.logger.setLevel(logging.DEBUG)
        self.stream_handler = logging.StreamHandler()
        self.logger.addHandler(self.stream_handler)
        self.file_handler = logging.FileHandler(log_address, mode='w')
        self.logger.addHandler(self.file_handler)
        self.logger.info('EXE INFO::' + str(datetime.datetime.now()))
        self.option_to_naver_id = {0:0, 1:2}
        self.option_to_kakao_id = {0:'SHORTEST_DIST', 1:'SHORTEST_REALTIME'}

    def setValue(self, config, _file_address, target, _sheet_name, _column1, _column2, option ,_startrow, _endrow):
        self.startrow = _startrow
        self.endrow = _endrow
        self.operation = self.startrow
        self.file_address = _file_address
        self.sheet_name = _sheet_name
        self.column1 = _column1
        self.column2 = _column2
        self.wb = excel.load_workbook(_file_address)
        self.sheet = self.wb[self.sheet_name]
        self.max_row = self.sheet.max_row
        self.percent = 0
        self.option = option
        self.config_list = config
        self.destination = target

    def request_to_naver(self, url):
        http_header = {'accept':'application/json, text/javascript, */*; q=0.01',  'accept-encoding':'gzip, deflate, br',
             'accept-language':'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
             'referer':'https://m.map.naver.com/directions/',
             'sec-fetch-mode':'cors',
             'sec-fetch-site':'same-origin',
             'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.88 Safari/537.36',
             'x-requested-with':'XMLHttpRequest'}

        r = requests.get(url, headers=http_header)
        return r

    def request_to_kakao(self, url, search=False):
        if search:
            http_header = {'Accept':'text/javascript, application/javascript, application/ecmascript, application/x-ecmascript, */*; q=0.01',
             'Accept-Encoding':'gzip, deflate, br',
             'Accept-Language':'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
             'Connection':'keep-alive',
             'Host':'map.kakao.com',
             'Referer':'https://map.kakao.com/',
             'Sec-Fetch-Mode':'cors',
             'Sec-Fetch-Site':'same-origin',
             'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.88 Safari/537.36',
             'X-Requested-With':'XMLHttpRequest'}
        else:
            http_header = {'Accept':'*/*',
             'Accept-Encoding':'gzip, deflate, br',
             'Accept-Language':'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
             'Connection':'keep-alive',
             'Host':'search.map.daum.net',
             'Referer':'https://map.kakao.com/',
             'Sec-Fetch-Mode':'no-cors',
             'Sec-Fetch-Site':'cross-site',
             'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.88 Safari/537.36'}

        r = requests.get(url, headers=http_header)
        return r    

    def naver_get_point_url(self, query):
        query = urllib.parse.quote_plus(query)
        url = 'https://m.map.naver.com/apis/search/poi?query={}&page=1'.format(query)
        return url

    def naver_get_distance_url(self, start_point, end_point):
        start_point = urllib.parse.quote_plus(start_point)
        end_point = urllib.parse.quote_plus(end_point)
        url = 'https://m.map.naver.com/spirra/findCarRoute.nhn?route=route3&output=json&result=web3&coord_type=latlng&search={}&car=0&mileage=12.4&start={}&destination={}'.format(self.option_to_naver_id[self.option], start_point, end_point)
        return url

    def naver_get_point(self, query, return_name=False):
        r = self.request_to_naver(self.naver_get_point_url(query))
        r.raise_for_status()
        res_dict = r.json()
        try:
            x = res_dict['result']['address']['list'][0]['x']
            y = res_dict['result']['address']['list'][0]['y']
            name = res_dict['result']['address']['list'][0]['name']
        except TypeError:
            x = res_dict['result']['site']['list'][0]['x']
            y = res_dict['result']['site']['list'][0]['y']
            name = res_dict['result']['site']['list'][0]['name']
        except:
            self.logger.error(('[Code P] 400 Bad Request or 404 Not Found'+query+'\n'+ r.text), exc_info=True)
            raise ValueError
        if return_name:
            return '{},{},{}'.format(x, y, name), name
        else:
            return '{},{},{}'.format(x, y, name)

    def kakao_get_point_url(self, query):
        query = urllib.parse.quote_plus(query)
        url = 'https://search.map.daum.net/mapsearch/map.daum?callback=jQuery181029877381355741384_1577093231120&q={}&msFlag=A&sort=0&gb=R'.format(query)
        return url

    def kakao_get_distance_url(self, start_point, start_id, end_point, end_id):
        start_point = urllib.parse.quote_plus(start_point)
        end_point = urllib.parse.quote_plus(end_point)
        url = 'https://map.kakao.com/route/carset.json?roadside=ON&sp={},ADDRESS,{}&ep={},ADDRESS,{}&callback=jQuery181029877381355741384_1577093231121&carMode={}&carOption=NONE'.format(start_point, start_point, end_point, end_id, self.option_to_kakao_id(self.option))
        return url

    def kakao_get_point(self, query, return_name=False):
        r = self.request_to_kakao(self.kakao_get_point_url(query), True)
        r.raise_for_status()
        res = r.text
        res = res[43:-3]
        res_dict = json.loads(res)
        try:
            x = res_dict['address'][0]['x']
            y = res_dict['address'][0]['y']
            docid = res_dict['address'][0]['docid']
            name = res_dict['address'][0]['addr']
        except:
            self.logger.error(('[Code P] 400 Bad Request or 404 Not Found\n' + res), exc_info=True)
            raise ValueError
        
        if return_name:
            return ('{},{},{}'.format(x, y, name), docid, name)
        else:
            return ('{},{},{}'.format(x, y, name), docid)

    def distance_query(self, query, platform):
        error = False
        end_point = self.destination_point
        if platform == 0:
            try:
                start_point = self.naver_get_point(query)
            except ValueError:
                self.logger.warning(('[Code Q1] Input Address not Exist::' + query), exc_info=True)
                distance = 0
                error = True
                return (error, distance)
            except:
                self.logger.error('[Code Q1] 400 Bad Request\n'+query, exc_info=True)
                raise PermissionError

            try:
                r = self.request_to_naver(self.naver_get_distance_url(start_point, end_point))
                r.raise_for_status()
                res = r.text
                res_dict = json.loads(res)
                target = res_dict['routes'][0]['summary']
                distance = target['distance']
            except Exception:
                self.logger.error(('[Code Q2] 400 Bad Request\n' + res), exc_info=True)
                raise PermissionError

        else:
            end_id = self.destination_id
            try:
                start_point, start_id = self.kakao_get_point(query)
            except ValueError:
                self.logger.warning(('[Code Q1] Input Address not Exist::' + query), exc_info=True)
                distance = 0
                error = True
                return (
                error, distance)
            except:
                self.logger.error('[Code Q1] 400 Bad Request\n', exc_info=True)
                raise PermissionError

            try:
                r = self.request_to_kakao(self.kakao_get_distance_url(start_point, start_id, end_point, end_id))
                r.raise_for_status()
                res = r.text
                res = res[42:-1]
                res_dict = json.loads(res)
                distance = res_dict['list'][0]['totalLength']
            except Exception:
                self.logger.error(('[Code Q2] 400 Bad Request\n' + res), exc_info=True)
                raise PermissionError

        return (error, distance)

    def load_query(self):
        start = None
        error = False
        if self.operation <= self.max_row:
            start = self.sheet[(self.column1 + str(self.operation))].value
            if start == None:
                start = 'N/A'
                error = True
            return (error, start)

    def save_result(self, error, distance):
        if self.operation <= self.max_row:
            if not error:
                self.sheet[self.column2 + str(self.operation)] = distance / 1000
                self.wb.save(self.file_address)
            else:
                self.sheet[self.column2 + str(self.operation)] = 'N/A'
                self.wb.save(self.file_address)
            return

    def fetch(self):
        self.total = self.endrow - self.startrow + 1
        last_percent = 0
        try:
            if self.config_list[0] == 0:
                self.destination_point, name = self.naver_get_point(self.destination, True)
            else:
                self.destination_point, self.destination_id, name = self.kakao_get_point(self.destination, True)
        except ValueError:
            self.permission.emit(3)
            return
        for i in range(self.total):
            if self.online:
                distance = 0
                rand = random.uniform(self.config_list[1], self.config_list[2])
                time.sleep(rand)
                if not self.online:
                    print('call break')
                    break
                if not self.config_list[3] == 0:
                    if (i + 1) % self.config_list[3] == 0:
                        time.sleep(self.config_list[4] * 60)
                st = time.time()
                if i == 0:
                    self.percentChanged.emit(0)
                try:
                    load_error, starting_point = self.load_query()
                except TypeError:
                    load_error = True
                except Exception as e:
                    self.logger.warning((str(e)), exc_info=True)

                if not load_error:
                    try:
                        fetch_error, distance = self.distance_query(starting_point, self.config_list[0])
                    except Exception as e:
                        self.logger.critical(('[Code F1] Network Error::' + str(e)), exc_info=True)
                        if self.config_list[5] == 0:
                            self.permission.emit(1)
                            return
                        for j in range(self.config_list[6], 0, -1):
                            self.waitsignal.emit(j)
                            time.sleep(60)
                        i -= 1
                        continue

                    try:
                        self.save_result(fetch_error, distance)
                    except Exception as e:
                        self.logger.critical(('[Code F2] Permission Denied::' + str(e)), exc_info=True)
                        self.permission.emit(2)
                        return

                    self.operation += 1
                    self.cur = i + 1
                    self.percent = self.cur / self.total * 100
                    et = time.time()
                    run_time = et - st
                    last_item = self.total - self.cur
                    last_time = (run_time + (self.config_list[1] + self.config_list[2]) / 2) * last_item + last_item // self.config_list[3] * self.config_list[4] * 60
                    emit_data = [self.total, starting_point, distance / 1000, last_time, last_item, self.cur, name]
                    if last_percent == 0:
                        self.completion.emit(emit_data)
                    for i in range(last_percent, int(self.percent) + 1):
                        self.percentChanged.emit(i)
                        time.sleep(0.05)

                    last_percent = int(self.percent)
                    if not last_percent == 0:
                        self.completion.emit(emit_data)
            else:
                self.logger.info('예기치 않게 Thread가 종료되었습니다.')
                break

        self.logger.info('Thread가 종료되었습니다.')
        self.wb.close()